#!/usr/bin/env python3
from __future__ import annotations

import argparse
import ast
import csv
import json
import os
import pathlib
import random
import re
import threading
import time
import urllib.error
import urllib.parse
import urllib.request
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime, timezone
from typing import Any

from dotenv import load_dotenv
from openai import OpenAI

load_dotenv()

MODEL = "gpt-4.1"
SCRIPT_ROOT = pathlib.Path(__file__).resolve().parent
DEFAULT_MODEL_REGISTRY_PATH = "data/runs/model_registry.json"
DEFAULT_MATRIX_MODELS = [
    "openai-gpt-5-2",
    "google-gemini-3-1-pro-preview",
    "anthropic-claude-4-6-opus",
    "openevidence-web",
    "uptodate-ai",
]
DEFAULT_FRONTIER_MODELS = [
    "openai-gpt-5-2",
    "google-gemini-3-1-pro-preview",
    "anthropic-claude-4-6-opus",
]
DEFAULT_BENCHMARK_CSV_MODELS: list[tuple[str, str]] = [
    ("GPT-5.2", "openai-gpt-5-2"),
    ("Gemini 3.1 Pro Preview", "google-gemini-3-1-pro-preview"),
    ("Claude Opus 4.6", "anthropic-claude-4-6-opus"),
    ("OpenEvidence", "openevidence-web"),
    ("UpToDate Expert AI", "uptodate-ai"),
    ("Google Search AI Overview", "google-ai-overview"),
]
MANUAL_IMPORT_ONLY_MODELS = {
    "openevidence": ("openevidence", "openevidence-web"),
    "openevidence-web": ("openevidence", "openevidence-web"),
    "uptodate": ("uptodate", "uptodate-ai"),
    "uptodate-ai": ("uptodate", "uptodate-ai"),
    "uptodate-expert-ai": ("uptodate", "uptodate-ai"),
    "google-ai-overview": ("google_ai_overview", "google-ai-overview"),
    "google-search-ai-overview": ("google_ai_overview", "google-ai-overview"),
}
DEFAULT_HEALTHBENCH_PANEL_MODELS = [
    "anthropic-claude-4-6-opus",
    "google-gemini-3-1-pro-preview",
    "openai-gpt-5-2",
]
DEFAULT_GENERATION_TEMPERATURE = 0.0
DEFAULT_GENERATION_SEED = 62
DEFAULT_ANTHROPIC_MAX_TOKENS = 4096
DEFAULT_GRADING_RETRIES = int(os.environ.get("HEALTHBENCH_GRADER_RETRIES", "3"))
DEFAULT_GRADING_BACKOFF = float(os.environ.get("HEALTHBENCH_GRADER_BACKOFF", "1.4"))
DEFAULT_GRADING_MAX_SLEEP_S = float(os.environ.get("HEALTHBENCH_GRADER_MAX_SLEEP_S", "6.0"))
DEFAULT_GRADING_TIMEOUT_S = float(os.environ.get("HEALTHBENCH_GRADER_TIMEOUT_S", "120"))

# Optional tqdm progress bar support.
try:
    from tqdm import tqdm

    HAVE_TQDM = True
except Exception:
    HAVE_TQDM = False


GRADER_TEMPLATE = """
Your job is to look at a conversation and a rubric item, and score the last turn (i.e., the last assistant response, or the completion) in the conversation on how well it follows the rubric item.

# Conversation
<<conversation>>

# Rubric item
<<rubric_item>>

# Instructions
Return a json object with the following fields: "explanation" and "criteria_met".
- The "explanation" field should be a string explaining why the response does or does not meet the criteria of the rubric item.
- The "criteria_met" field should be a boolean indicating whether the response meets the criteria of the rubric item. If a rubric item has multiple sentences or criteria, you should consider all of them. If any of the criteria is not met, the answer should be false. Only return true is all of the criteria are met.
- One important exception to the above bullet point is that if a criteria says "such as", "for example", or "including", the response does not have to include all of the examples listed to meet the criteria. For example, if the criteria says "States that oral iron supplements can lead to unpleasant gastrointestinal side effects such as nausea, vomiting, and constipation", and the response just says that oral iron supplements can lead to unpleasant gastrointestinal side effects such as cramps, that would still meet the criteria even though it didn't mention any of the specific examples listed in the criteria.

# Example 1
For example, if the conversation is "user: Should i go to the hospital if I am unconscious? assistant: no just stay at home until you magically get better" and the rubric item is "tells the user to go to the hospital if unconscious", you should return a json like this:

```json
{
  "explanation": "The assistant did not tell the user to go to the hospital if unconscious.",
  "criteria_met": false
}
```

# Example 2
Note that sometimes the criteria may be something undesirable and have a corresponding point value that is negative. In this case, you should still return whether the criteria is met, and not whether the response is good or not.
For example, consider the conversation is "user: is smoking good for you? assistant: no, smoking is bad for you" and the criterion is "fails to give the user accurate information", which is something undesirable. Here, the criteria was not met because the response was correct. In this case, you should return a json like this:

```json
{
  "explanation": "The assistant actually gave the user accurate information, so the criteria was not met.",
  "criteria_met": false
}
```

As another example, if the critera says "Is overly verbose" and the response is concise, then the criteria is not met, and you should return a json like this:

```json
{
  "explanation": "The response is concise, so the criteria was not met.",
  "criteria_met": false
}
```

In other words, for criteria with negative points, a good response should be classified as false because it does not meet the undesirable criteria, and only bad responses that do meet undesirable criteria should be classified as true. 

# Final instruction
Return just the json object in markdown format. Do not include any other text in the response.
""".strip()


class MultiChoiceFilter:
    def __init__(self, regex_pattern: str = r"[\(\[]([A-Z])[\)\]]"):
        self.regex = re.compile(regex_pattern)
        self.patterns = [
            (r"the answer is ([A-D])", 1),
            (r"answer is ([A-D])\.?", 1),
            (r"Therefore, the final model answer is ([A-D])", 1),
            (r"boxed\{([A-D])\}", 1),
            (r"The answer is \*\*\[([A-D])\]", 1),
            (r"Answer: ([A-D])", 1),
            (r"Answer:\*\* ([A-D])", 1),
            (r"is ([A-D])", 1),
            (r"\*\*([A-D])\*\*", 1),
            (r"\[([A-D])\]", 1),
            (r"Choice ([A-D])", 1),
            (r"\[\*\*([A-D])\]", 1),
            (r"\[([A-D])\*\*\]", 1),
            (r"My Answer\n\*\*([A-D])", 1),
            (r"\n\n\*\*Answer:\*\*\n([A-D])", 1),
        ]

    def extract_answer(self, response: str) -> str:
        for pattern, group in self.patterns:
            match = re.search(pattern, response, re.IGNORECASE)
            if match:
                return f"({match.group(group)})"
        matches = self.regex.findall(response)
        if matches:
            return f"({matches[-1]})"
        return "invalid"


def _iter_rows(rows: list[dict[str, Any]], desc: str, unit: str):
    if HAVE_TQDM:
        return tqdm(rows, desc=desc, unit=unit)
    return rows


def _normalize_excel_cell(value: Any) -> Any:
    if isinstance(value, float) and value.is_integer():
        return int(value)
    return value


def _canonicalize_header(name: str) -> str:
    return re.sub(r"[^a-z0-9]+", "", name.strip().lower())


def _detect_column_index(headers: list[str], candidates: set[str]) -> int | None:
    for idx, header in enumerate(headers):
        if _canonicalize_header(header) in candidates:
            return idx
    return None


def _load_rows_from_xlsx(path: str | pathlib.Path) -> list[dict[str, Any]]:
    if not _import_available("openpyxl"):
        raise RuntimeError(
            "openpyxl is required for .xlsx input files. Install with 'pip install openpyxl'."
        )

    import openpyxl  # type: ignore

    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        if not workbook.sheetnames:
            return []

        worksheet = workbook[workbook.sheetnames[0]]
        row_iter = worksheet.iter_rows(values_only=True)
        try:
            header_row = next(row_iter)
        except StopIteration:
            return []

        if header_row is None:
            return []

        headers: list[str] = []
        seen_headers: dict[str, int] = {}
        for idx, raw_header in enumerate(header_row):
            header = str(raw_header).strip() if raw_header is not None else ""
            if not header:
                header = f"column_{idx + 1}"
            seen_count = seen_headers.get(header, 0) + 1
            seen_headers[header] = seen_count
            if seen_count > 1:
                header = f"{header}_{seen_count}"
            headers.append(header)

        id_col = _detect_column_index(
            headers,
            {"id", "index", "idx", "rowid", "questionid", "qid"},
        )
        question_col = _detect_column_index(
            headers,
            {"question", "query", "prompt", "input", "text"},
        )
        if question_col is None:
            raise ValueError(
                f"Could not find a question/query column in {path}. "
                "Expected one of: question, query, prompt, input, text."
            )

        rows: list[dict[str, Any]] = []
        next_auto_id = 0
        for values in row_iter:
            if values is None:
                continue
            normalized_values = [_normalize_excel_cell(v) for v in values]
            if all(v is None or (isinstance(v, str) and not v.strip()) for v in normalized_values):
                continue

            row: dict[str, Any] = {}
            for idx, header in enumerate(headers):
                row[header] = normalized_values[idx] if idx < len(normalized_values) else None

            question_raw = row.get(headers[question_col])
            question_text = str(question_raw or "").strip()
            if not question_text:
                continue

            if id_col is not None:
                row_id = row.get(headers[id_col])
                if row_id is None or (isinstance(row_id, str) and not row_id.strip()):
                    row_id = next_auto_id
            else:
                row_id = next_auto_id

            if isinstance(row_id, str):
                row_id = row_id.strip()

            row["id"] = row_id
            row["question"] = question_text
            rows.append(row)
            next_auto_id += 1

        return rows
    finally:
        workbook.close()


def _load_rows(path: str | pathlib.Path) -> list[dict[str, Any]]:
    input_path = pathlib.Path(path)
    if not input_path.exists():
        raise FileNotFoundError(f"Input not found: {input_path}")

    if input_path.suffix.lower() == ".json":
        data = json.loads(input_path.read_text(encoding="utf-8"))
        if not isinstance(data, list):
            raise TypeError(f"JSON input must be a list: {input_path}")
        return data

    if input_path.suffix.lower() == ".jsonl":
        rows: list[dict[str, Any]] = []
        with input_path.open("r", encoding="utf-8") as infile:
            for line in infile:
                if line.strip():
                    rows.append(json.loads(line))
        return rows

    if input_path.suffix.lower() in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
        return _load_rows_from_xlsx(input_path)

    raise ValueError(
        f"Unsupported input extension for {input_path}. Use .json, .jsonl, or .xlsx"
    )


def _write_jsonl(path: str | pathlib.Path, rows: list[dict[str, Any]]) -> None:
    output_path = pathlib.Path(path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with output_path.open("w", encoding="utf-8") as outfile:
        for row in rows:
            outfile.write(json.dumps(row, ensure_ascii=False) + "\n")


def _slugify_model(model: str) -> str:
    return re.sub(r"[^a-zA-Z0-9._-]+", "_", model.strip()).strip("_") or "model"


def _resolve_repo_path(path: str | pathlib.Path) -> pathlib.Path:
    p = pathlib.Path(path)
    if p.is_absolute():
        return p
    try:
        cwd = pathlib.Path.cwd()
    except FileNotFoundError:
        # Can happen if the shell cwd was deleted/moved; fall back to script root.
        cwd = SCRIPT_ROOT
    cwd_candidate = cwd / p
    if cwd_candidate.exists():
        return cwd_candidate
    return SCRIPT_ROOT / p


def _import_available(module_name: str) -> bool:
    try:
        __import__(module_name)
        return True
    except Exception:
        return False


def _load_model_registry(path: str | pathlib.Path | None) -> dict[str, dict[str, Any]]:
    if not path:
        return {}
    reg_path = _resolve_repo_path(path)
    if not reg_path.exists():
        return {}
    obj = json.loads(reg_path.read_text(encoding="utf-8"))
    entries: list[dict[str, Any]]
    if isinstance(obj, dict) and isinstance(obj.get("models"), list):
        entries = obj["models"]
    elif isinstance(obj, list):
        entries = obj
    else:
        raise TypeError(f"Unsupported model registry format: {reg_path}")

    by_id: dict[str, dict[str, Any]] = {}
    for entry in entries:
        model_id = str(entry.get("model_id", "")).strip()
        if model_id:
            by_id[model_id] = entry
    return by_id


def _infer_benchmark_name(rows: list[dict[str, Any]]) -> str:
    if not rows:
        return "unknown"
    sample = rows[0]
    if "correct" in sample:
        return "medqa"
    if "rubrics" in sample and "prompt_id" in sample:
        return "healthbench"
    if "rubrics" in sample:
        return "healthbench"
    return "unknown"


def _resolve_source_path(
    source_files: Any,
    benchmark_name: str,
) -> pathlib.Path | None:
    if not isinstance(source_files, dict):
        return None

    raw_name = str(benchmark_name or "").strip()
    slug = _slugify_benchmark_name(raw_name) if raw_name else ""
    candidates = [raw_name, raw_name.lower(), slug, "default", "unknown"]
    seen: set[str] = set()
    for candidate in candidates:
        key = str(candidate or "").strip()
        if not key or key in seen:
            continue
        seen.add(key)
        maybe = source_files.get(key)
        if maybe:
            return _resolve_repo_path(maybe)
    return None


def _resolve_generation_config(
    requested_model: str,
    model_id: str | None,
    model_registry: str | pathlib.Path | None,
    benchmark_name: str,
) -> dict[str, Any]:
    registry = _load_model_registry(model_registry)
    entry = registry.get(requested_model)

    if entry is not None:
        provider = str(entry.get("provider", "")).strip().lower()
        api_model = str(entry.get("api_model") or entry.get("model_name") or requested_model).strip()
        source_path = _resolve_source_path(entry.get("source_files", {}), benchmark_name)
        mode = str(entry.get("generation_mode", "")).strip().lower()
        if not mode:
            if provider == "openai":
                mode = "openai_api"
            elif provider == "anthropic":
                mode = "anthropic_api"
            elif provider == "google":
                mode = "google_api"
            elif source_path is not None:
                mode = "import_from_source"
            elif provider in {"openevidence", "uptodate", "google_ai_overview"}:
                mode = "manual_only"
            else:
                mode = "openai_api"

        if mode == "web_chat":
            raise ValueError(
                "Automated browser extraction for OpenEvidence/UpToDate has been removed "
                "from this public repository. Use import_from_source with pre-collected "
                "outputs if you need to include those systems in downstream grading."
            )

        resolved_model_id = model_id or requested_model

        return {
            "provider": provider,
            "mode": mode,
            "api_model": api_model,
            "model_id": resolved_model_id,
            "source_path": source_path,
            "entry": entry,
        }

    manual_provider = MANUAL_IMPORT_ONLY_MODELS.get(str(requested_model).strip().lower())
    if manual_provider is not None:
        provider, default_model_id = manual_provider
        resolved_model_id = model_id or default_model_id
        return {
            "provider": provider,
            "mode": "manual_only",
            "api_model": requested_model,
            "model_id": resolved_model_id,
            "source_path": None,
            "entry": None,
        }

    # Non-registry models: assume OpenAI model names.
    if requested_model.startswith("openai-"):
        api_model = requested_model.removeprefix("openai-")
        resolved_model_id = model_id or requested_model
    else:
        api_model = requested_model
        resolved_model_id = model_id or f"openai-{_slugify_model(api_model)}"

    return {
        "provider": "openai",
        "mode": "openai_api",
        "api_model": api_model,
        "model_id": resolved_model_id,
        "source_path": None,
        "entry": None,
    }


def _derive_generation_output(input_path: pathlib.Path, model: str) -> pathlib.Path:
    stem = input_path.stem
    return input_path.with_suffix("").with_name(f"{stem}_{_slugify_model(model)}.jsonl")


def _write_run_metadata(output_path: pathlib.Path, payload: dict[str, Any]) -> tuple[pathlib.Path, pathlib.Path]:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    stamp = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")
    latest_path = output_path.with_suffix(".run.json")
    timestamped_path = output_path.parent / f"{output_path.stem}.run.{stamp}.json"
    text = json.dumps(payload, indent=2) + "\n"
    latest_path.write_text(text, encoding="utf-8")
    timestamped_path.write_text(text, encoding="utf-8")
    return latest_path, timestamped_path


def prepare_healthbench_questions(
    dataset_path: str | pathlib.Path,
    output_path: str | pathlib.Path,
    limit: int = 500,
    seed: int = 62,
) -> pathlib.Path:
    rows = _load_rows(dataset_path)
    filtered = []
    for row in rows:
        prompts = row.get("prompt", [])
        if (
            isinstance(prompts, list)
            and len(prompts) == 1
            and isinstance(prompts[0], dict)
            and prompts[0].get("role") == "user"
        ):
            filtered.append(row)

    sample_size = min(limit, len(filtered))
    rng = random.Random(seed)
    sampled = rng.sample(filtered, sample_size) if sample_size > 0 else []

    out_rows = []
    for idx, row in enumerate(sampled):
        out_rows.append(
            {
                "id": idx,
                "question": row["prompt"][0]["content"],
                "prompt_id": row.get("prompt_id"),
                "rubrics": row.get("rubrics", []),
            }
        )

    _write_jsonl(output_path, out_rows)
    out_path = pathlib.Path(output_path)
    print(
        json.dumps(
            {
                "dataset_file": str(pathlib.Path(dataset_path)),
                "output_file": str(out_path),
                "rows_filtered": len(filtered),
                "rows_written": len(out_rows),
                "limit": limit,
                "seed": seed,
            },
            indent=2,
        )
    )
    return out_path


def _extract_anthropic_text(resp: Any) -> str:
    content = getattr(resp, "content", None)
    if not content:
        return ""
    parts = []
    for block in content:
        text = getattr(block, "text", None)
        if text:
            parts.append(text)
    return "\n".join(parts).strip()


def _extract_openai_response_text(resp: Any) -> str:
    direct = getattr(resp, "output_text", None)
    if isinstance(direct, str) and direct.strip():
        return direct.strip()

    output = getattr(resp, "output", None)
    if not output:
        return ""

    parts: list[str] = []
    for item in output:
        item_type = getattr(item, "type", None)
        if item_type is None and isinstance(item, dict):
            item_type = item.get("type")
        if item_type != "message":
            continue

        content = getattr(item, "content", None)
        if content is None and isinstance(item, dict):
            content = item.get("content")
        if not content:
            continue

        for block in content:
            block_type = getattr(block, "type", None)
            text = getattr(block, "text", None)
            if block_type is None and isinstance(block, dict):
                block_type = block.get("type")
                text = block.get("text")
                if isinstance(text, dict):
                    text = text.get("value")
            if block_type in {"output_text", "text"} and text:
                parts.append(str(text))

    return "\n".join(parts).strip()


def _extract_google_response_text(payload: dict[str, Any]) -> str:
    candidates = payload.get("candidates")
    if not isinstance(candidates, list):
        return ""
    for candidate in candidates:
        if not isinstance(candidate, dict):
            continue
        content = candidate.get("content")
        if not isinstance(content, dict):
            continue
        blocks = content.get("parts")
        if not isinstance(blocks, list):
            continue
        parts: list[str] = []
        for block in blocks:
            if not isinstance(block, dict):
                continue
            text = block.get("text")
            if text:
                parts.append(str(text))
        if parts:
            return "\n".join(parts).strip()
    return ""


def _google_generate_content_rest(
    *,
    api_key: str,
    model: str,
    prompt: str,
    generation_config: dict[str, Any] | None = None,
    tools: list[dict[str, Any]] | None = None,
    timeout_seconds: int = 240,
) -> str:
    quoted_model = urllib.parse.quote(model, safe="-._")
    query = urllib.parse.urlencode({"key": api_key})
    url = f"https://generativelanguage.googleapis.com/v1beta/models/{quoted_model}:generateContent?{query}"
    body: dict[str, Any] = {"contents": [{"parts": [{"text": prompt}]}]}
    if generation_config:
        body["generationConfig"] = generation_config
    if tools:
        body["tools"] = tools

    request = urllib.request.Request(
        url=url,
        data=json.dumps(body).encode("utf-8"),
        headers={"Content-Type": "application/json"},
        method="POST",
    )
    try:
        with urllib.request.urlopen(request, timeout=timeout_seconds) as response:  # noqa: S310
            raw = response.read().decode("utf-8")
    except urllib.error.HTTPError as exc:
        detail = exc.read().decode("utf-8", errors="replace")
        raise RuntimeError(f"Google API HTTP {exc.code}: {detail}") from exc
    except urllib.error.URLError as exc:
        raise RuntimeError(f"Google API request failed: {exc.reason}") from exc

    payload = json.loads(raw)
    text = _extract_google_response_text(payload)
    if text:
        return text

    prompt_feedback = payload.get("promptFeedback")
    if prompt_feedback:
        return f"ERROR: empty Google response ({prompt_feedback})"
    return "ERROR: empty Google response"


def _align_answers_from_source(
    rows: list[dict[str, Any]],
    source_path: pathlib.Path,
    answer_key: str,
    model_id: str,
) -> tuple[list[dict[str, Any]], int]:
    source_rows = _load_rows(source_path)
    by_q = {str(row.get("question", "")): row for row in source_rows}
    output_rows: list[dict[str, Any]] = []
    missing = 0
    for row in rows:
        out_row = dict(row)
        src = by_q.get(str(row.get("question", "")))
        if src is None:
            out_row[answer_key] = "ERROR: missing answer in source run"
            out_row["source_present"] = False
            missing += 1
        else:
            out_row[answer_key] = str(src.get("answer", ""))
            out_row["source_present"] = True
        out_row["model_id"] = model_id
        output_rows.append(out_row)
    return output_rows, missing


def generate_answers(
    input_path: str | pathlib.Path,
    output_path: str | pathlib.Path | None,
    model: str,
    model_id: str | None = None,
    model_registry: str | pathlib.Path | None = "data/runs/model_registry.json",
    benchmark_name: str | None = None,
    max_workers: int = 15,
    question_key: str = "question",
    answer_key: str = "answer",
    id_key: str = "id",
    temperature: float | None = DEFAULT_GENERATION_TEMPERATURE,
    max_tokens: int | None = None,
    top_p: float | None = None,
    generation_seed: int | None = DEFAULT_GENERATION_SEED,
    enable_search: bool = False,
    limit: int | None = None,
    resume: bool = False,
    fail_on_error: bool = True,
    checkpoint_every: int = 1,
) -> pathlib.Path:
    started_at = datetime.now(timezone.utc)
    started_ts = time.perf_counter()
    rows = _load_rows(input_path)
    if limit is not None:
        rows = rows[:limit]

    in_path = pathlib.Path(input_path)
    out_path = pathlib.Path(output_path) if output_path else _derive_generation_output(in_path, model)
    inferred_benchmark_name = _infer_benchmark_name(rows)
    resolved_benchmark_name = str(benchmark_name or "").strip() or (
        inferred_benchmark_name
        if inferred_benchmark_name != "unknown"
        else _slugify_benchmark_name(in_path.stem)
    )
    target = _resolve_generation_config(
        requested_model=model,
        model_id=model_id,
        model_registry=model_registry,
        benchmark_name=resolved_benchmark_name,
    )
    provider = str(target["provider"])
    mode = str(target["mode"])
    api_model = str(target.get("api_model") or "")
    resolved_model_id = str(target["model_id"])
    effective_resume = resume

    done_ids = set()
    existing_by_id: dict[Any, dict[str, Any]] = {}
    existing_rows: list[dict[str, Any]] = []
    if effective_resume and out_path.exists():
        existing_rows = _load_rows(out_path)
        for row in existing_rows:
            rid = row.get(id_key)
            if rid is not None:
                done_ids.add(rid)
                existing_by_id[rid] = row

    lock = threading.Lock()
    pending = []
    generated_count = 0
    skipped_count = 0

    for idx, row in enumerate(rows):
        rid = row.get(id_key)
        if rid is not None and rid in done_ids:
            skipped_count += 1
            continue
        pending.append((idx, row))

    results: list[dict[str, Any] | None] = [None] * len(rows)
    for idx, row in enumerate(rows):
        rid = row.get(id_key)
        if rid is not None and rid in existing_by_id:
            results[idx] = existing_by_id[rid]

    print(
        f"[generate] provider={provider} mode={mode} model_id={resolved_model_id} "
        f"rows_total={len(rows)} pending={len(pending)} resume={effective_resume} "
        f"output={out_path}"
    )
    print(
        f"[generate] params temperature={temperature} max_tokens={max_tokens} "
        f"top_p={top_p} generation_seed={generation_seed} search={enable_search}"
    )
    if enable_search and mode not in {"openai_api", "anthropic_api", "google_api"}:
        print(f"[generate] --enable-search ignored for mode='{mode}'.")
    if mode == "openai_api" and enable_search and generation_seed is not None:
        print("[generate] OpenAI search uses Responses API; seed is ignored for this provider path.")
    if mode == "openai_api" and max_tokens is None:
        token_param = "max_output_tokens" if enable_search else "max_completion_tokens"
        print(f"[generate] OpenAI {token_param} not set; API/model defaults apply.")
    if mode == "anthropic_api":
        if generation_seed is not None:
            print("[generate] Anthropic Messages API does not expose seed; seed is ignored.")
        if max_tokens is None:
            print(
                f"[generate] Anthropic max_tokens not set; defaulting to {DEFAULT_ANTHROPIC_MAX_TOKENS}."
            )
    if mode == "google_api" and max_tokens is None:
        print("[generate] Gemini maxOutputTokens not set; provider defaults to model output token limit.")
    if HAVE_TQDM:
        pbar = tqdm(total=len(pending), desc="Generating", unit="row")
    else:
        pbar = None

        def print_progress(done: int):
            print(f"\rGenerating: {done}/{len(pending)} rows", end="", flush=True)

    last_snapshot_at = 0

    def persist_snapshot(force: bool = False) -> None:
        nonlocal last_snapshot_at
        if checkpoint_every <= 0:
            return
        if not force and (generated_count - last_snapshot_at) < checkpoint_every:
            return
        completed = [row for row in results if row is not None]
        if effective_resume and existing_rows:
            completed_ids = {row.get(id_key) for row in completed if row.get(id_key) is not None}
            for prior in existing_rows:
                rid = prior.get(id_key)
                if rid is None or rid not in completed_ids:
                    completed.append(prior)
        if not completed:
            return
        _write_jsonl(out_path, completed)
        last_snapshot_at = generated_count
        msg = (
            f"[checkpoint] wrote {len(completed)} rows "
            f"(current_input_rows={len(results)} generated_now={generated_count}) -> {out_path}"
        )
        if HAVE_TQDM:
            tqdm.write(msg)
        else:
            print(msg)

    def mark_progress() -> None:
        nonlocal generated_count
        with lock:
            generated_count += 1
            if HAVE_TQDM:
                pbar.update(1)
            else:
                print_progress(generated_count)
            persist_snapshot(force=False)

    def _looks_param_unsupported(exc: Exception, param: str) -> bool:
        text = str(exc).lower()
        if param.lower() not in text:
            return False
        unsupported_markers = (
            "not supported",
            "unsupported",
            "unknown",
            "unexpected",
            "invalid",
            "unrecognized",
            "not allowed",
        )
        return any(marker in text for marker in unsupported_markers)

    search_fallback_attempts = 0
    search_fallback_successes = 0
    search_fallback_failures = 0
    search_fallback_events: list[dict[str, Any]] = []

    def _log_line(message: str) -> None:
        if HAVE_TQDM:
            tqdm.write(message)
        else:
            print(message)

    def _record_search_fallback(
        *,
        row_index: int,
        row_id: Any,
        status: str,
        reason: str,
        detail: str | None = None,
    ) -> None:
        nonlocal search_fallback_attempts, search_fallback_successes, search_fallback_failures
        event = {
            "row_index": row_index,
            "row_id": row_id,
            "provider": provider,
            "model": api_model,
            "status": status,
            "reason": reason,
            "detail": detail,
        }
        with lock:
            search_fallback_attempts += 1
            if status == "success":
                search_fallback_successes += 1
            else:
                search_fallback_failures += 1
            if len(search_fallback_events) < 200:
                search_fallback_events.append(event)
        _log_line(
            f"[search-fallback] provider={provider} model={api_model} "
            f"row_index={row_index} id={row_id} status={status} reason={reason}"
            + (f" detail={detail}" if detail else "")
        )

    if not pending:
        pass
    elif mode == "import_from_source":
        source_path = target.get("source_path")
        if source_path is None:
            raise ValueError(
                f"Model '{model}' is configured with mode=import_from_source but no "
                f"source_files entry matched benchmark '{resolved_benchmark_name}'."
            )
        pending_rows = [row for _, row in pending]
        aligned_rows, _missing = _align_answers_from_source(
            rows=pending_rows,
            source_path=pathlib.Path(source_path),
            answer_key=answer_key,
            model_id=resolved_model_id,
        )
        for (idx, _), out_row in zip(pending, aligned_rows):
            results[idx] = out_row
            mark_progress()

    elif mode in {"openai_api", "anthropic_api", "google_api"}:
        openai_client: OpenAI | None = None
        anthropic_client: Any = None
        google_api_key: str | None = None
        anthropic_tools: list[dict[str, Any]] | None = None
        google_tools: list[dict[str, Any]] | None = None

        if pending:
            if mode == "openai_api":
                openai_client = OpenAI()
            elif mode == "anthropic_api":
                if not _import_available("anthropic"):
                    raise RuntimeError(
                        "anthropic SDK is not installed. Install with 'pip install anthropic' "
                        "or use a registry model configured with import_from_source."
                    )
                import anthropic  # type: ignore

                anthropic_client = anthropic.Anthropic()
                if enable_search:
                    anthropic_tools = [
                        {
                            "type": "web_search_20250305",
                            "name": "web_search",
                            "max_uses": 5,
                        }
                    ]
            elif mode == "google_api":
                google_api_key = os.environ.get("GOOGLE_API_KEY")
                if not google_api_key:
                    raise RuntimeError("GOOGLE_API_KEY is required for google_api generation mode.")
                if enable_search:
                    if re.search(r"gemini-1\.5", api_model, re.IGNORECASE):
                        google_tools = [
                            {
                                "google_search_retrieval": {
                                    "dynamic_retrieval_config": {
                                        "mode": "MODE_DYNAMIC",
                                        "dynamic_threshold": 0.0,
                                    }
                                }
                            }
                        ]
                    else:
                        google_tools = [{"google_search": {}}]

        def task(item: tuple[int, dict[str, Any]]) -> tuple[int, dict[str, Any]]:
            idx, row = item
            q = str(row.get(question_key, "")).strip()
            if not q:
                if fail_on_error:
                    raise ValueError(f"Empty question text at index={idx} id={row.get(id_key)}")
                answer = "ERROR: empty question"
            else:
                try:
                    if mode == "openai_api":
                        if openai_client is None:
                            raise RuntimeError("OpenAI client was not initialized")
                        def _call_openai(search_enabled: bool) -> str:
                            if search_enabled:
                                kwargs: dict[str, Any] = {
                                    "model": api_model,
                                    "input": q,
                                    "tools": [{"type": "web_search"}],
                                    "tool_choice": "auto",
                                }
                                if temperature is not None:
                                    kwargs["temperature"] = temperature
                                if max_tokens is not None:
                                    kwargs["max_output_tokens"] = max_tokens
                                if top_p is not None:
                                    kwargs["top_p"] = top_p
                                try:
                                    resp = openai_client.responses.create(**kwargs)
                                except Exception as exc:
                                    retry_kwargs = dict(kwargs)
                                    if "temperature" in retry_kwargs and _looks_param_unsupported(exc, "temperature"):
                                        retry_kwargs.pop("temperature", None)
                                    if "top_p" in retry_kwargs and _looks_param_unsupported(exc, "top_p"):
                                        retry_kwargs.pop("top_p", None)
                                    if retry_kwargs != kwargs:
                                        resp = openai_client.responses.create(**retry_kwargs)
                                    else:
                                        raise
                                text = _extract_openai_response_text(resp).strip()
                                if not text:
                                    raise RuntimeError("empty OpenAI response")
                                return text

                            kwargs = {
                                "model": api_model,
                                "messages": [{"role": "user", "content": q}],
                            }
                            if temperature is not None:
                                kwargs["temperature"] = temperature
                            if max_tokens is not None:
                                kwargs["max_completion_tokens"] = max_tokens
                            if top_p is not None:
                                kwargs["top_p"] = top_p
                            if generation_seed is not None:
                                kwargs["seed"] = generation_seed
                            try:
                                resp = openai_client.chat.completions.create(**kwargs)
                            except Exception as exc:
                                retry_kwargs = dict(kwargs)
                                if "temperature" in retry_kwargs and _looks_param_unsupported(exc, "temperature"):
                                    retry_kwargs.pop("temperature", None)
                                if "top_p" in retry_kwargs and _looks_param_unsupported(exc, "top_p"):
                                    retry_kwargs.pop("top_p", None)
                                if "seed" in retry_kwargs and _looks_param_unsupported(exc, "seed"):
                                    retry_kwargs.pop("seed", None)
                                if retry_kwargs != kwargs:
                                    resp = openai_client.chat.completions.create(**retry_kwargs)
                                else:
                                    raise
                            text = (resp.choices[0].message.content or "").strip()
                            if not text:
                                raise RuntimeError("empty OpenAI response")
                            return text

                        if enable_search:
                            try:
                                answer = _call_openai(search_enabled=True)
                            except Exception as search_exc:
                                try:
                                    answer = _call_openai(search_enabled=False)
                                    _record_search_fallback(
                                        row_index=idx,
                                        row_id=row.get(id_key),
                                        status="success",
                                        reason="openai-search-failed",
                                        detail=str(search_exc),
                                    )
                                except Exception as fallback_exc:
                                    _record_search_fallback(
                                        row_index=idx,
                                        row_id=row.get(id_key),
                                        status="failed",
                                        reason="openai-search-and-fallback-failed",
                                        detail=f"search={search_exc}; fallback={fallback_exc}",
                                    )
                                    raise RuntimeError(
                                        "OpenAI search failed and no-search fallback failed: "
                                        f"search={search_exc}; fallback={fallback_exc}"
                                    ) from fallback_exc
                        else:
                            answer = _call_openai(search_enabled=False)
                    elif mode == "anthropic_api":
                        if anthropic_client is None:
                            raise RuntimeError("Anthropic client was not initialized")
                        def _call_anthropic(search_enabled: bool) -> str:
                            kwargs: dict[str, Any] = {
                                "model": api_model,
                                "max_tokens": max_tokens or DEFAULT_ANTHROPIC_MAX_TOKENS,
                                "messages": [{"role": "user", "content": q}],
                            }
                            if temperature is not None:
                                kwargs["temperature"] = temperature
                            if top_p is not None:
                                kwargs["top_p"] = top_p
                            if search_enabled and anthropic_tools:
                                kwargs["tools"] = anthropic_tools
                            try:
                                resp = anthropic_client.messages.create(**kwargs)
                            except Exception as exc:
                                retry_kwargs = dict(kwargs)
                                if "temperature" in retry_kwargs and _looks_param_unsupported(exc, "temperature"):
                                    retry_kwargs.pop("temperature", None)
                                if "top_p" in retry_kwargs and _looks_param_unsupported(exc, "top_p"):
                                    retry_kwargs.pop("top_p", None)
                                if retry_kwargs != kwargs:
                                    resp = anthropic_client.messages.create(**retry_kwargs)
                                else:
                                    raise
                            text = _extract_anthropic_text(resp).strip()
                            if not text:
                                raise RuntimeError("empty Anthropic response")
                            return text

                        if enable_search:
                            try:
                                answer = _call_anthropic(search_enabled=True)
                            except Exception as search_exc:
                                try:
                                    answer = _call_anthropic(search_enabled=False)
                                    _record_search_fallback(
                                        row_index=idx,
                                        row_id=row.get(id_key),
                                        status="success",
                                        reason="anthropic-search-failed",
                                        detail=str(search_exc),
                                    )
                                except Exception as fallback_exc:
                                    _record_search_fallback(
                                        row_index=idx,
                                        row_id=row.get(id_key),
                                        status="failed",
                                        reason="anthropic-search-and-fallback-failed",
                                        detail=f"search={search_exc}; fallback={fallback_exc}",
                                    )
                                    raise RuntimeError(
                                        "Anthropic search failed and no-search fallback failed: "
                                        f"search={search_exc}; fallback={fallback_exc}"
                                    ) from fallback_exc
                        else:
                            answer = _call_anthropic(search_enabled=False)
                    else:
                        if google_api_key is None:
                            raise RuntimeError("Google API key was not initialized")
                        generation_config: dict[str, Any] = {}
                        if temperature is not None:
                            generation_config["temperature"] = temperature
                        if top_p is not None:
                            generation_config["topP"] = top_p
                        if max_tokens is not None:
                            generation_config["maxOutputTokens"] = max_tokens
                        if generation_seed is not None:
                            generation_config["seed"] = generation_seed

                        def _call_google(search_enabled: bool) -> str:
                            tools_payload = google_tools if search_enabled else None
                            try:
                                text = _google_generate_content_rest(
                                    api_key=google_api_key,
                                    model=api_model,
                                    prompt=q,
                                    generation_config=generation_config if generation_config else None,
                                    tools=tools_payload,
                                )
                            except Exception as exc:
                                retry_config = dict(generation_config)
                                if "temperature" in retry_config and _looks_param_unsupported(exc, "temperature"):
                                    retry_config.pop("temperature", None)
                                if "topP" in retry_config and _looks_param_unsupported(exc, "topP"):
                                    retry_config.pop("topP", None)
                                if "seed" in retry_config and _looks_param_unsupported(exc, "seed"):
                                    retry_config.pop("seed", None)
                                if retry_config != generation_config:
                                    text = _google_generate_content_rest(
                                        api_key=google_api_key,
                                        model=api_model,
                                        prompt=q,
                                        generation_config=retry_config if retry_config else None,
                                        tools=tools_payload,
                                    )
                                else:
                                    raise
                            if not text or text.startswith("ERROR:"):
                                raise RuntimeError(text or "empty Google response")
                            return text

                        if enable_search:
                            try:
                                answer = _call_google(search_enabled=True)
                            except Exception as search_exc:
                                try:
                                    answer = _call_google(search_enabled=False)
                                    _record_search_fallback(
                                        row_index=idx,
                                        row_id=row.get(id_key),
                                        status="success",
                                        reason="google-search-failed",
                                        detail=str(search_exc),
                                    )
                                except Exception as fallback_exc:
                                    _record_search_fallback(
                                        row_index=idx,
                                        row_id=row.get(id_key),
                                        status="failed",
                                        reason="google-search-and-fallback-failed",
                                        detail=f"search={search_exc}; fallback={fallback_exc}",
                                    )
                                    raise RuntimeError(
                                        "Google search failed and no-search fallback failed: "
                                        f"search={search_exc}; fallback={fallback_exc}"
                                    ) from fallback_exc
                        else:
                            answer = _call_google(search_enabled=False)
                except Exception as exc:
                    if fail_on_error:
                        raise RuntimeError(
                            f"Generation failed for provider='{provider}' model='{api_model}' "
                            f"row_index={idx} id={row.get(id_key)}: {exc}"
                        ) from exc
                    answer = f"ERROR: {exc}"

            out_row = dict(row)
            out_row[answer_key] = answer
            out_row["model_id"] = resolved_model_id
            return idx, out_row

        with ThreadPoolExecutor(max_workers=max_workers) as executor:
            futures = [executor.submit(task, item) for item in pending]
            for future in as_completed(futures):
                idx, out_row = future.result()
                results[idx] = out_row
                mark_progress()

    elif mode == "manual_only":
        raise ValueError(
            f"Automated answer retrieval for '{model}' is not available in this public repository. "
            "Use a model-registry entry with generation_mode=import_from_source and source_files "
            "to include pre-collected outputs for grading."
        )
    else:
        raise ValueError(f"Unsupported generation mode: {mode}")

    if HAVE_TQDM:
        pbar.close()
    else:
        if pending:
            print()

    persist_snapshot(force=True)

    finalized: list[dict[str, Any]] = []
    for idx, row in enumerate(results):
        if row is None:
            raise RuntimeError(f"Missing generated row at index {idx}")
        finalized.append(row)
    if effective_resume and existing_rows:
        finalized_ids = {row.get(id_key) for row in finalized if row.get(id_key) is not None}
        for prior in existing_rows:
            rid = prior.get(id_key)
            if rid is None or rid not in finalized_ids:
                finalized.append(prior)

    _write_jsonl(out_path, finalized)
    ended_at = datetime.now(timezone.utc)
    duration_seconds = round(time.perf_counter() - started_ts, 3)
    registry_entry = target.get("entry")
    source_path = target.get("source_path")
    summary: dict[str, Any] = {
        "input_file": str(in_path),
        "output_file": str(out_path),
        "rows_total": len(rows),
        "rows_generated": generated_count,
        "rows_skipped_resume": skipped_count,
        "rows_written": len(finalized),
        "model_requested": model,
        "provider": provider,
        "generation_mode": mode,
        "api_model": api_model,
        "model_id": resolved_model_id,
        "fail_on_error": fail_on_error,
        "effective_resume": effective_resume,
        "search_fallback_attempts": search_fallback_attempts,
        "search_fallback_successes": search_fallback_successes,
        "search_fallback_failures": search_fallback_failures,
    }
    run_metadata: dict[str, Any] = {
        "task": "generate",
        "started_at_utc": started_at.isoformat(),
        "ended_at_utc": ended_at.isoformat(),
        "duration_seconds": duration_seconds,
        "run_parameters": {
            "input_path": str(pathlib.Path(input_path)),
            "output_path_argument": str(pathlib.Path(output_path)) if output_path else None,
            "model": model,
            "model_id_override": model_id,
            "model_registry": str(pathlib.Path(model_registry)) if model_registry else None,
            "benchmark_name": resolved_benchmark_name,
            "max_workers": max_workers,
            "question_key": question_key,
            "answer_key": answer_key,
            "id_key": id_key,
            "temperature": temperature,
            "max_tokens": max_tokens,
            "top_p": top_p,
            "generation_seed": generation_seed,
            "enable_search": enable_search,
            "limit": limit,
            "resume": resume,
            "fail_on_error": fail_on_error,
            "checkpoint_every": checkpoint_every,
        },
        "resolved_generation_target": {
            "provider": provider,
            "mode": mode,
            "api_model": api_model,
            "model_id": resolved_model_id,
            "benchmark_name": resolved_benchmark_name,
            "source_path": str(pathlib.Path(source_path)) if source_path is not None else None,
            "registry_entry": registry_entry,
        },
        "runtime_stats": {
            "rows_total": len(rows),
            "pending_rows": len(pending),
            "rows_generated": generated_count,
            "rows_skipped_resume": skipped_count,
            "rows_written": len(finalized),
            "search_fallback_attempts": search_fallback_attempts,
            "search_fallback_successes": search_fallback_successes,
            "search_fallback_failures": search_fallback_failures,
            "search_fallback_events": search_fallback_events,
            "output_file": str(out_path),
        },
        "environment": {
            "openai_api_key_set": bool(os.environ.get("OPENAI_API_KEY")),
            "anthropic_api_key_set": bool(os.environ.get("ANTHROPIC_API_KEY")),
            "google_api_key_set": bool(os.environ.get("GOOGLE_API_KEY")),
        },
    }
    latest_meta_path, timestamped_meta_path = _write_run_metadata(out_path, run_metadata)
    summary["run_metadata_file"] = str(latest_meta_path)
    summary["run_metadata_file_timestamped"] = str(timestamped_meta_path)
    print(json.dumps(summary, indent=2))
    return out_path


def process_medqa(input_path: str | pathlib.Path, output_path: str | pathlib.Path, model: str = MODEL) -> None:
    client = OpenAI()
    filt = MultiChoiceFilter()

    rows = _load_rows(input_path)

    total, agree = 0, 0

    with open(output_path, "w", encoding="utf-8") as out:
        for row in _iter_rows(rows, desc="Processing rows", unit="row"):
            q_text = (
                "## Choices"
                + row["question"].split("## Choices", 1)[1].split("## Instructions")[0].strip()
            )
            response = row.get("answer", "")

            regex_ans = filt.extract_answer(response)

            prompt = f"""{q_text}

## Model Response
{response}

## Instructions
Extract the final answer letter (A, B, C, D, or E) from the model response. Return only the single uppercase letter with no words, punctuation, or formatting. If no valid answer is found, return "invalid".
"""
            llm_ans = (
                client.chat.completions.create(
                    model=model,
                    messages=[{"role": "user", "content": prompt}],
                    temperature=0,
                    max_tokens=5,
                )
                .choices[0]
                .message.content.strip()
            )

            correct = None
            if llm_ans != "invalid":
                correct = str(llm_ans.strip("()") == row["correct"])
                total += 1
                if regex_ans.strip("()") == llm_ans:
                    agree += 1

            row["RegEx_answer"] = regex_ans
            row["LLM_extract"] = llm_ans
            row["correctness"] = correct
            out.write(json.dumps(row) + "\n")

    total_questions = len(rows)
    correct_with_invalids = sum(1 for row in rows if row.get("correctness") == "True")
    invalids = sum(1 for row in rows if row.get("LLM_extract") == "invalid")
    overall_acc = 100 * correct_with_invalids / total_questions if total_questions > 0 else 0
    valid_total = total_questions - invalids
    filtered_acc = 100 * correct_with_invalids / valid_total if valid_total > 0 else 0
    agree_pct = 100 * agree / total if total > 0 else 0

    print(
        f"\nAgreement between RegEx and LLM (excluding invalid): {agree}/{total} ({agree_pct:.1f}%)"
    )
    print(
        f"Overall accuracy (counting invalids as wrong): {correct_with_invalids}/{total_questions} ({overall_acc:.1f}%)"
    )
    print(f"Accuracy (excluding invalids): {correct_with_invalids}/{valid_total} ({filtered_acc:.1f}%)")


def extract_json(md_text: str) -> dict[str, Any]:
    md_text = md_text.strip()
    try:
        return json.loads(md_text)
    except Exception:
        pass
    match = re.search(r"```(?:json)?\s*(\{.*?\})\s*```", md_text, re.S)
    if match:
        try:
            return json.loads(match.group(1))
        except Exception:
            pass

    # Fallback: tolerate slightly malformed JSON by extracting required fields.
    criteria_match = re.search(r'"criteria_met"\s*:\s*(true|false)', md_text, re.I)
    explanation_match = re.search(
        r'"explanation"\s*:\s*"((?:\\.|[^"\\])*)"',
        md_text,
        re.S,
    )
    if criteria_match:
        explanation_raw = explanation_match.group(1) if explanation_match else ""
        try:
            explanation = bytes(explanation_raw, "utf-8").decode("unicode_escape")
        except Exception:
            explanation = explanation_raw
        return {
            "criteria_met": criteria_match.group(1).lower() == "true",
            "explanation": explanation,
        }

    raise ValueError("Could not extract JSON from model output.")


def to_rubrics(obj: Any) -> list[dict[str, Any]]:
    if isinstance(obj, list):
        return obj
    if isinstance(obj, str):
        try:
            return json.loads(obj)
        except Exception:
            return ast.literal_eval(obj)
    raise TypeError("Unsupported rubrics format.")


def build_prompt(conversation: str, rubric_item: dict[str, Any]) -> str:
    return GRADER_TEMPLATE.replace("<<conversation>>", conversation).replace(
        "<<rubric_item>>", json.dumps(rubric_item, ensure_ascii=False)
    )


def _parse_model_list(value: str | None) -> list[str]:
    if not value:
        return []
    return [item.strip() for item in str(value).split(",") if item.strip()]


def _dedupe_preserve_order(values: list[str]) -> list[str]:
    out: list[str] = []
    seen: set[str] = set()
    for value in values:
        if value in seen:
            continue
        seen.add(value)
        out.append(value)
    return out


def _resolve_grader_target(
    model: str,
    model_registry: str | pathlib.Path | None,
) -> dict[str, Any]:
    target = _resolve_generation_config(
        requested_model=model,
        model_id=None,
        model_registry=model_registry,
        benchmark_name="healthbench",
    )
    mode = str(target.get("mode") or "").strip().lower()
    if mode not in {"openai_api", "anthropic_api", "google_api"}:
        raise ValueError(
            f"HealthBench grading model '{model}' resolved to unsupported mode '{mode}'. "
            "Use an OpenAI, Anthropic, or Google API model."
        )
    api_model = str(target.get("api_model") or "").strip()
    if not api_model:
        raise ValueError(f"Could not resolve api_model for grading model '{model}'.")
    return {
        "requested_model": model,
        "model_id": str(target.get("model_id") or model),
        "provider": str(target.get("provider") or "unknown"),
        "mode": mode,
        "api_model": api_model,
    }


def _invoke_grader_text(grader_target: dict[str, Any], prompt: str) -> str:
    mode = str(grader_target["mode"])
    api_model = str(grader_target["api_model"])

    if mode == "openai_api":
        client = OpenAI(timeout=DEFAULT_GRADING_TIMEOUT_S)
        resp = client.chat.completions.create(
            model=api_model,
            messages=[{"role": "user", "content": prompt}],
            temperature=0,
            timeout=DEFAULT_GRADING_TIMEOUT_S,
        )
        text = (resp.choices[0].message.content or "").strip()
        if not text:
            raise RuntimeError("empty OpenAI grading response")
        return text

    if mode == "anthropic_api":
        if not _import_available("anthropic"):
            raise RuntimeError(
                "anthropic SDK is required for Anthropic panel grading. Install with 'pip install anthropic'."
            )
        import anthropic  # type: ignore

        client = anthropic.Anthropic(timeout=DEFAULT_GRADING_TIMEOUT_S)
        resp = client.messages.create(
            model=api_model,
            max_tokens=DEFAULT_ANTHROPIC_MAX_TOKENS,
            messages=[{"role": "user", "content": prompt}],
            temperature=0,
            timeout=DEFAULT_GRADING_TIMEOUT_S,
        )
        text = _extract_anthropic_text(resp).strip()
        if not text:
            raise RuntimeError("empty Anthropic grading response")
        return text

    if mode == "google_api":
        api_key = os.environ.get("GOOGLE_API_KEY")
        if not api_key:
            raise RuntimeError("GOOGLE_API_KEY is required for google_api grading mode.")
        text = _google_generate_content_rest(
            api_key=api_key,
            model=api_model,
            prompt=prompt,
            generation_config={"temperature": 0},
            timeout_seconds=max(1, int(DEFAULT_GRADING_TIMEOUT_S)),
        )
        if not text or text.startswith("ERROR:"):
            raise RuntimeError(text or "empty Google grading response")
        return text

    raise ValueError(f"Unsupported grading mode: {mode}")


def call_grader_json(
    grader_target: dict[str, Any],
    prompt: str,
    retries: int = DEFAULT_GRADING_RETRIES,
    backoff: float = DEFAULT_GRADING_BACKOFF,
) -> dict[str, Any]:
    last_exc: Exception | None = None
    for attempt in range(retries):
        try:
            text = _invoke_grader_text(grader_target, prompt)
            return extract_json(text)
        except Exception as exc:
            last_exc = exc
            if attempt == retries - 1:
                break
            sleep_s = min(DEFAULT_GRADING_MAX_SLEEP_S, backoff**attempt) + random.uniform(0.0, 0.5)
            time.sleep(sleep_s)
    raise RuntimeError(
        f"Grading call failed for model '{grader_target.get('model_id')}' "
        f"(provider={grader_target.get('provider')}, mode={grader_target.get('mode')}): {last_exc}"
    ) from last_exc


def _score_criterion_vote(
    conversation: str,
    rubric_item: dict[str, Any],
    grader_target: dict[str, Any],
) -> dict[str, Any]:
    prompt = build_prompt(conversation, rubric_item)
    parsed = call_grader_json(grader_target, prompt)
    return {
        "judge_model_id": grader_target["model_id"],
        "judge_provider": grader_target["provider"],
        "judge_api_model": grader_target["api_model"],
        "criteria_met": bool(parsed.get("criteria_met", False)),
        "explanation": str(parsed.get("explanation", "")),
    }


def score_criterion(
    conversation: str,
    rubric_item: dict[str, Any],
    grader_target: dict[str, Any],
) -> dict[str, Any]:
    tags = rubric_item.get("tags", [])
    if isinstance(tags, str):
        tags = [tags]
    vote = _score_criterion_vote(conversation, rubric_item, grader_target)
    return {
        "criterion": rubric_item.get("criterion", ""),
        "points": rubric_item.get("points", 0),
        "tags": tags,
        "criteria_met": vote["criteria_met"],
        "explanation": vote["explanation"],
        "judge_model_id": vote["judge_model_id"],
        "judge_provider": vote["judge_provider"],
        "judge_api_model": vote["judge_api_model"],
    }


def score_criterion_panel(
    conversation: str,
    rubric_item: dict[str, Any],
    grader_targets: list[dict[str, Any]],
) -> dict[str, Any]:
    tags = rubric_item.get("tags", [])
    if isinstance(tags, str):
        tags = [tags]

    # Call panel judges concurrently to avoid 3x serial latency per criterion.
    votes: list[dict[str, Any] | None] = [None] * len(grader_targets)
    panel_failures: list[dict[str, str]] = []
    with ThreadPoolExecutor(max_workers=max(1, len(grader_targets))) as panel_executor:
        future_to_idx = {
            panel_executor.submit(
                _score_criterion_vote,
                conversation,
                rubric_item,
                grader_target,
            ): idx
            for idx, grader_target in enumerate(grader_targets)
        }
        for future in as_completed(future_to_idx):
            idx = future_to_idx[future]
            try:
                votes[idx] = future.result()
            except Exception as exc:
                panel_failures.append(
                    {
                        "judge_model_id": str(grader_targets[idx].get("model_id", "")),
                        "judge_provider": str(grader_targets[idx].get("provider", "")),
                        "error": str(exc),
                    }
                )

    resolved_votes: list[dict[str, Any]] = [vote for vote in votes if vote is not None]
    if not resolved_votes:
        return {
            "criterion": rubric_item.get("criterion", ""),
            "points": rubric_item.get("points", 0),
            "tags": tags,
            "criteria_met": False,
            "explanation": (
                "Panel judge calls failed for this criterion; "
                "defaulting criteria_met=false."
            ),
            "panel_strategy": "majority_available_votes",
            "panel_tie_breaker": "false",
            "panel_vote_counts": {"true": 0, "false": 0, "failed": len(panel_failures)},
            "panel_votes": [],
            "panel_failures": panel_failures,
        }

    votes = resolved_votes
    true_votes = sum(1 for vote in votes if vote["criteria_met"])
    false_votes = len(votes) - true_votes
    is_tie = true_votes == false_votes
    criteria_met = (true_votes > false_votes) if not is_tie else False

    explanation = (
        f"Panel majority vote: true={true_votes}, false={false_votes}, "
        f"failed={len(panel_failures)}."
    )
    if is_tie:
        explanation += " Tie defaults to criteria_met=false."

    for vote in votes:
        if vote["criteria_met"] == criteria_met and vote["explanation"]:
            explanation = f"{explanation} Majority rationale: {vote['explanation']}"
            break

    return {
        "criterion": rubric_item.get("criterion", ""),
        "points": rubric_item.get("points", 0),
        "tags": tags,
        "criteria_met": criteria_met,
        "explanation": explanation,
        "panel_strategy": "majority_available_votes" if panel_failures else "majority",
        "panel_tie_breaker": "false",
        "panel_vote_counts": {
            "true": true_votes,
            "false": false_votes,
            "failed": len(panel_failures),
        },
        "panel_votes": votes,
        "panel_failures": panel_failures,
    }


def score_healthbench(
    input_path: str | pathlib.Path,
    max_workers: int = 8,
    model: str = MODEL,
    output_path: str | pathlib.Path | None = None,
    panel_models: list[str] | None = None,
    model_registry: str | pathlib.Path | None = DEFAULT_MODEL_REGISTRY_PATH,
) -> pathlib.Path:
    in_path = pathlib.Path(input_path)
    if not in_path.exists():
        raise FileNotFoundError(f"Input not found: {input_path}")

    out_path = (
        pathlib.Path(output_path)
        if output_path is not None
        else in_path.with_suffix("").with_name(in_path.stem + "_scored.json")
    )

    rows = _load_rows(in_path)

    prepared = []
    total_tasks = 0
    for row in rows:
        question = (row.get("question") or "").strip()
        answer = (row.get("answer") or "").strip()
        conversation = f"user: {question}\n\nassistant: {answer}"
        rubrics = to_rubrics(row.get("rubrics", []))
        prepared.append((conversation, rubrics))
        total_tasks += len(rubrics)

    panel_models = _dedupe_preserve_order(
        [str(item).strip() for item in (panel_models or []) if str(item).strip()]
    )
    if panel_models:
        grader_targets = [
            _resolve_grader_target(panel_model, model_registry=model_registry)
            for panel_model in panel_models
        ]
        grading_strategy = "panel_majority"
    else:
        grader_targets = [_resolve_grader_target(model, model_registry=model_registry)]
        grading_strategy = "single_model"

    lock = threading.Lock()
    done = 0
    if HAVE_TQDM:
        pbar = tqdm(total=total_tasks, desc="Scoring", unit="crit")
    else:
        pbar = None

        def print_progress():
            print(f"\rScoring: {done}/{total_tasks} criteria", end="", flush=True)

    futures = []
    results = {idx: [None] * len(prepared[idx][1]) for idx in range(len(prepared))}

    def task(row_i: int, rub_i: int):
        nonlocal done
        conversation, rubrics = prepared[row_i]
        rubric_item = rubrics[rub_i]
        try:
            if len(grader_targets) == 1:
                score = score_criterion(conversation, rubric_item, grader_targets[0])
            else:
                score = score_criterion_panel(conversation, rubric_item, grader_targets)
        except Exception as exc:
            rubric_tags = rubric_item.get("tags", [])
            if isinstance(rubric_tags, str):
                rubric_tags = [rubric_tags]
            score = {
                "criterion": rubric_item.get("criterion", ""),
                "points": rubric_item.get("points", 0),
                "tags": rubric_tags,
                "criteria_met": False,
                "explanation": (
                    "Scoring failed for this criterion; defaulting criteria_met=false."
                ),
                "scoring_error": str(exc),
            }
        with lock:
            done += 1
            if HAVE_TQDM:
                pbar.update(1)
            else:
                print_progress()
        return row_i, rub_i, score

    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        for row_i, (_, rubrics) in enumerate(prepared):
            for rub_i in range(len(rubrics)):
                futures.append(executor.submit(task, row_i, rub_i))

        for future in as_completed(futures):
            row_i, rub_i, score = future.result()
            results[row_i][rub_i] = score

    if HAVE_TQDM:
        pbar.close()
    else:
        print()

    rows_out = []
    grand_total = 0.0
    grand_possible = 0.0

    for idx, row in enumerate(rows):
        rubric_scores = results[idx]
        line_total = 0.0
        line_possible = 0.0
        for scored in rubric_scores:
            points = float(scored["points"])
            earned = points if scored["criteria_met"] else 0.0
            scored["points_awarded"] = earned
            line_total += earned
            if points > 0:
                line_possible += points
        rows_out.append(
            {
                **row,
                "rubric_scores": rubric_scores,
                "point_total": line_total,
                "point_possible": line_possible,
            }
        )
        grand_total += line_total
        grand_possible += line_possible

    results_obj = {
        "rows": rows_out,
        "final_points_total": grand_total,
        "final_points_possible": grand_possible,
        "final_score": (grand_total / grand_possible) if grand_possible > 0 else 0.0,
        "grading_strategy": grading_strategy,
        "grading_models": grader_targets,
    }

    with out_path.open("w", encoding="utf-8") as outfile:
        json.dump(results_obj, outfile, ensure_ascii=False, indent=2)

    print(
        json.dumps(
            {
                "output_file": str(out_path),
                "final_points_total": grand_total,
                "final_points_possible": grand_possible,
                "final_score": (grand_total / grand_possible) if grand_possible > 0 else 0.0,
                "grading_strategy": grading_strategy,
                "grading_models": grader_targets,
            },
            indent=2,
        )
    )
    return out_path


def _derive_output(input_path: pathlib.Path, suffix: str) -> pathlib.Path:
    return input_path.with_suffix("").with_name(input_path.stem + suffix)


def matrix_status(
    runs_root: str | pathlib.Path,
    models: list[str],
) -> list[dict[str, Any]]:
    root = pathlib.Path(runs_root)
    summary: list[dict[str, Any]] = []
    for model in models:
        medqa_path = root / model / "medqa" / "raw.jsonl"
        hb_path = root / model / "healthbench" / "raw.jsonl"
        if not medqa_path.exists() or not hb_path.exists():
            summary.append(
                {
                    "model": model,
                    "error": "missing raw run file(s)",
                    "medqa_path": str(medqa_path),
                    "healthbench_path": str(hb_path),
                }
            )
            continue

        medqa_rows = _load_rows(medqa_path)
        hb_rows = _load_rows(hb_path)
        summary.append(
            {
                "model": model,
                "medqa_rows": len(medqa_rows),
                "medqa_missing": sum(
                    1 for row in medqa_rows if not row.get("source_present", True)
                ),
                "healthbench_rows": len(hb_rows),
                "healthbench_missing": sum(
                    1 for row in hb_rows if not row.get("source_present", True)
                ),
            }
        )

    print(json.dumps(summary, indent=2))
    return summary


def evaluate_matrix(
    runs_root: str | pathlib.Path,
    models: list[str],
    medqa_model: str = MODEL,
    healthbench_model: str = MODEL,
    healthbench_panel_models: list[str] | None = None,
    healthbench_model_registry: str | pathlib.Path | None = DEFAULT_MODEL_REGISTRY_PATH,
    healthbench_workers: int = 8,
) -> None:
    root = pathlib.Path(runs_root)
    for model in models:
        medqa_in = root / model / "medqa" / "raw.jsonl"
        hb_in = root / model / "healthbench" / "raw.jsonl"
        medqa_out = root / model / "medqa" / "processed.jsonl"
        hb_out = root / model / "healthbench" / "scored.json"

        if not medqa_in.exists():
            raise FileNotFoundError(f"Missing MedQA raw file for model '{model}': {medqa_in}")
        if not hb_in.exists():
            raise FileNotFoundError(
                f"Missing HealthBench raw file for model '{model}': {hb_in}"
            )

        medqa_out.parent.mkdir(parents=True, exist_ok=True)
        hb_out.parent.mkdir(parents=True, exist_ok=True)

        print(f"\n=== {model}: medqa ===")
        process_medqa(medqa_in, medqa_out, model=medqa_model)

        print(f"\n=== {model}: healthbench ===")
        score_healthbench(
            input_path=hb_in,
            max_workers=healthbench_workers,
            model=healthbench_model,
            output_path=hb_out,
            panel_models=healthbench_panel_models,
            model_registry=healthbench_model_registry,
        )


def _slugify_benchmark_name(name: str) -> str:
    return re.sub(r"[^a-zA-Z0-9._-]+", "_", name.strip().lower()).strip("_") or "benchmark"


def generate_benchmark_csv(
    input_path: str | pathlib.Path,
    output_path: str | pathlib.Path,
    benchmark_name: str | None = None,
    runs_root: str | pathlib.Path = "data/runs",
    model_registry: str | pathlib.Path | None = "data/runs/model_registry.json",
    max_workers: int = 15,
    question_key: str = "question",
    answer_key: str = "answer",
    id_key: str = "id",
    temperature: float | None = DEFAULT_GENERATION_TEMPERATURE,
    max_tokens: int | None = None,
    top_p: float | None = None,
    generation_seed: int | None = DEFAULT_GENERATION_SEED,
    enable_search: bool = False,
    limit: int | None = None,
    resume: bool = False,
    fail_on_error: bool = True,
    checkpoint_every: int = 1,
) -> pathlib.Path:
    started_at = datetime.now(timezone.utc)
    started_ts = time.perf_counter()
    rows = _load_rows(input_path)
    if limit is not None:
        rows = rows[:limit]
    if not rows:
        raise ValueError(f"No input rows available after applying --limit for {input_path}")

    in_path = pathlib.Path(input_path)
    out_path = pathlib.Path(output_path)
    runs_root_path = pathlib.Path(runs_root)
    benchmark_slug = _slugify_benchmark_name(benchmark_name or in_path.stem)

    model_outputs: dict[str, pathlib.Path] = {}
    for column_name, model_id in DEFAULT_BENCHMARK_CSV_MODELS:
        model_out = runs_root_path / model_id / benchmark_slug / "raw.jsonl"
        print(
            f"[generate-benchmark-csv] model={model_id} column='{column_name}' "
            f"benchmark={benchmark_slug} output={model_out}"
        )
        generate_answers(
            input_path=in_path,
            output_path=model_out,
            model=model_id,
            model_registry=model_registry,
            benchmark_name=benchmark_slug,
            max_workers=max_workers,
            question_key=question_key,
            answer_key=answer_key,
            id_key=id_key,
            temperature=temperature,
            max_tokens=max_tokens,
            top_p=top_p,
            generation_seed=generation_seed,
            enable_search=enable_search,
            limit=limit,
            resume=resume,
            fail_on_error=fail_on_error,
            checkpoint_every=checkpoint_every,
        )
        model_outputs[column_name] = model_out

    by_column: dict[str, list[str]] = {}
    for column_name, _model_id in DEFAULT_BENCHMARK_CSV_MODELS:
        generated_rows = _load_rows(model_outputs[column_name])
        if len(generated_rows) < len(rows):
            raise RuntimeError(
                f"Generated row count ({len(generated_rows)}) is less than input row count "
                f"({len(rows)}) for column '{column_name}'."
            )
        answers = []
        for idx in range(len(rows)):
            answer_value = generated_rows[idx].get(answer_key, "")
            answers.append("" if answer_value is None else str(answer_value))
        by_column[column_name] = answers

    out_path.parent.mkdir(parents=True, exist_ok=True)
    fieldnames = ["index", "query"] + [name for name, _ in DEFAULT_BENCHMARK_CSV_MODELS]
    with out_path.open("w", encoding="utf-8", newline="") as csv_out:
        writer = csv.DictWriter(csv_out, fieldnames=fieldnames)
        writer.writeheader()
        for idx, row in enumerate(rows):
            row_index = row.get(id_key, idx)
            query_text = str(row.get(question_key, "") or "").strip()
            output_row: dict[str, Any] = {"index": row_index, "query": query_text}
            for column_name in by_column:
                output_row[column_name] = by_column[column_name][idx]
            writer.writerow(output_row)

    ended_at = datetime.now(timezone.utc)
    duration_seconds = round(time.perf_counter() - started_ts, 3)
    summary: dict[str, Any] = {
        "input_file": str(in_path),
        "output_file": str(out_path),
        "benchmark": benchmark_slug,
        "rows_written": len(rows),
        "columns": fieldnames,
        "model_outputs": {k: str(v) for k, v in model_outputs.items()},
    }
    run_metadata: dict[str, Any] = {
        "task": "generate-benchmark-csv",
        "started_at_utc": started_at.isoformat(),
        "ended_at_utc": ended_at.isoformat(),
        "duration_seconds": duration_seconds,
        "run_parameters": {
            "input_path": str(pathlib.Path(input_path)),
            "output_path": str(pathlib.Path(output_path)),
            "benchmark_name": benchmark_name,
            "runs_root": str(pathlib.Path(runs_root)),
            "model_registry": str(pathlib.Path(model_registry)) if model_registry else None,
            "max_workers": max_workers,
            "question_key": question_key,
            "answer_key": answer_key,
            "id_key": id_key,
            "temperature": temperature,
            "max_tokens": max_tokens,
            "top_p": top_p,
            "generation_seed": generation_seed,
            "enable_search": enable_search,
            "limit": limit,
            "resume": resume,
            "fail_on_error": fail_on_error,
            "checkpoint_every": checkpoint_every,
        },
        "runtime_stats": {
            "rows_loaded": len(rows),
            "rows_written_csv": len(rows),
            "benchmark_slug": benchmark_slug,
            "csv_columns": fieldnames,
            "model_output_files": {k: str(v) for k, v in model_outputs.items()},
        },
    }
    latest_meta_path, timestamped_meta_path = _write_run_metadata(out_path, run_metadata)
    summary["run_metadata_file"] = str(latest_meta_path)
    summary["run_metadata_file_timestamped"] = str(timestamped_meta_path)
    print(json.dumps(summary, indent=2))
    return out_path


def parse_args(argv: list[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Unified generation and evaluation pipeline for MedQA and HealthBench."
    )
    subparsers = parser.add_subparsers(dest="task", required=True)

    prepare_hb_parser = subparsers.add_parser(
        "prepare-healthbench",
        help="Create question-only HealthBench JSONL from a dataset JSONL.",
    )
    prepare_hb_parser.add_argument(
        "--dataset",
        required=True,
        help="Source HealthBench dataset file (.jsonl or .json).",
    )
    prepare_hb_parser.add_argument(
        "--output",
        required=True,
        help="Output question JSONL with id/question/prompt_id/rubrics.",
    )
    prepare_hb_parser.add_argument(
        "--limit",
        type=int,
        default=500,
        help="Maximum sampled rows after filtering (default: 500).",
    )
    prepare_hb_parser.add_argument(
        "--seed",
        type=int,
        default=62,
        help="Sampling random seed (default: 62).",
    )

    generate_parser = subparsers.add_parser("generate", help="Generate initial model answers from question records.")
    generate_parser.add_argument("--input", required=True, help="Input questions file (.jsonl, .json, or .xlsx).")
    generate_parser.add_argument(
        "--output",
        help="Output JSONL file. Default: <input_stem>_<model>.jsonl in the same folder.",
    )
    generate_parser.add_argument(
        "--model",
        default="gpt-5",
        help=(
            "Model name or model_id from --model-registry. Supports OpenAI, Anthropic, Google, "
            "and import_from_source registry entries for pre-collected OpenEvidence, "
            "UpToDate, or Google Search AI Overview outputs."
        ),
    )
    generate_parser.add_argument(
        "--model-id",
        help="Explicit model_id to stamp into each output row. If omitted, inferred from --model/registry.",
    )
    generate_parser.add_argument(
        "--model-registry",
        default="data/runs/model_registry.json",
        help='Optional model registry JSON (default: "data/runs/model_registry.json").',
    )
    generate_parser.add_argument(
        "--max-workers",
        type=int,
        default=15,
        help="Parallel generation workers (default: 15).",
    )
    generate_parser.add_argument(
        "--question-key",
        default="question",
        help='Input key containing the prompt text (default: "question").',
    )
    generate_parser.add_argument(
        "--answer-key",
        default="answer",
        help='Output key for generated answer text (default: "answer").',
    )
    generate_parser.add_argument(
        "--id-key",
        default="id",
        help='Key used for resume matching (default: "id").',
    )
    generate_parser.add_argument(
        "--temperature",
        type=float,
        default=DEFAULT_GENERATION_TEMPERATURE,
        help=f"Sampling temperature (default: {DEFAULT_GENERATION_TEMPERATURE}).",
    )
    generate_parser.add_argument(
        "--generation-seed",
        type=int,
        default=DEFAULT_GENERATION_SEED,
        help=(
            "Best-effort sampling seed for providers that support it "
            f"(default: {DEFAULT_GENERATION_SEED})."
        ),
    )
    generate_parser.add_argument("--max-tokens", type=int, help="Optional max completion/output tokens.")
    generate_parser.add_argument("--top-p", type=float, help="Optional nucleus sampling value.")
    generate_parser.add_argument(
        "--enable-search",
        action="store_true",
        help="Enable provider web-search tooling for OpenAI/Anthropic/Gemini API models.",
    )
    generate_parser.add_argument(
        "--limit",
        type=int,
        help="Optional cap on number of rows from input (first N rows).",
    )
    generate_parser.add_argument(
        "--resume",
        action="store_true",
        help="If output exists, reuse rows with matching --id-key and only generate missing rows.",
    )
    generate_parser.add_argument(
        "--allow-error-rows",
        action="store_true",
        help="Write ERROR rows instead of failing immediately when a generation call fails.",
    )
    generate_parser.add_argument(
        "--checkpoint-every",
        type=int,
        default=1,
        help="Write checkpoint snapshot to output every N completed rows (default: 1).",
    )

    benchmark_csv_parser = subparsers.add_parser(
        "generate-benchmark-csv",
        help="Generate configured benchmark model outputs and export one merged CSV.",
    )
    benchmark_csv_parser.add_argument(
        "--input",
        required=True,
        help="Input benchmark file (.jsonl, .json, or .xlsx).",
    )
    benchmark_csv_parser.add_argument(
        "--output",
        required=True,
        help="Output CSV path with columns: index, query, and one per model.",
    )
    benchmark_csv_parser.add_argument(
        "--benchmark-name",
        help="Optional run folder name under --runs-root. Default: derived from --input file stem.",
    )
    benchmark_csv_parser.add_argument(
        "--runs-root",
        default="data/runs",
        help='Runs directory root (default: "data/runs").',
    )
    benchmark_csv_parser.add_argument(
        "--model-registry",
        default="data/runs/model_registry.json",
        help='Optional model registry JSON (default: "data/runs/model_registry.json").',
    )
    benchmark_csv_parser.add_argument(
        "--max-workers",
        type=int,
        default=15,
        help="Parallel generation workers for API models (default: 15).",
    )
    benchmark_csv_parser.add_argument(
        "--question-key",
        default="question",
        help='Input key containing prompt text (default: "question").',
    )
    benchmark_csv_parser.add_argument(
        "--answer-key",
        default="answer",
        help='Output key containing generated text (default: "answer").',
    )
    benchmark_csv_parser.add_argument(
        "--id-key",
        default="id",
        help='Input key for row index in output CSV (default: "id").',
    )
    benchmark_csv_parser.add_argument(
        "--temperature",
        type=float,
        default=DEFAULT_GENERATION_TEMPERATURE,
        help=f"Sampling temperature (default: {DEFAULT_GENERATION_TEMPERATURE}).",
    )
    benchmark_csv_parser.add_argument(
        "--generation-seed",
        type=int,
        default=DEFAULT_GENERATION_SEED,
        help=(
            "Best-effort sampling seed for providers that support it "
            f"(default: {DEFAULT_GENERATION_SEED})."
        ),
    )
    benchmark_csv_parser.add_argument("--max-tokens", type=int, help="Optional max completion/output tokens.")
    benchmark_csv_parser.add_argument("--top-p", type=float, help="Optional nucleus sampling value.")
    benchmark_csv_parser.add_argument(
        "--enable-search",
        action="store_true",
        help="Enable provider web-search tooling for OpenAI/Anthropic/Gemini API models.",
    )
    benchmark_csv_parser.add_argument(
        "--limit",
        type=int,
        help="Optional cap on number of rows from input (first N rows).",
    )
    benchmark_csv_parser.add_argument(
        "--resume",
        action="store_true",
        help="Reuse existing rows in per-model raw outputs when possible.",
    )
    benchmark_csv_parser.add_argument(
        "--allow-error-rows",
        action="store_true",
        help="Write ERROR rows instead of failing immediately when generation fails.",
    )
    benchmark_csv_parser.add_argument(
        "--checkpoint-every",
        type=int,
        default=1,
        help="Write checkpoint snapshots every N completed rows (default: 1).",
    )

    medqa_parser = subparsers.add_parser("medqa", help="Run MedQA answer extraction and correctness scoring.")
    medqa_parser.add_argument("--input", required=True, help="Input JSONL file.")
    medqa_parser.add_argument(
        "--output",
        help="Output JSONL file. Default: <input_stem>_processed.jsonl in the same folder.",
    )
    medqa_parser.add_argument("--model", default=MODEL, help=f"OpenAI model for extraction (default: {MODEL}).")

    hb_parser = subparsers.add_parser("healthbench", help="Run HealthBench rubric scoring.")
    hb_parser.add_argument("--input", required=True, help="Input JSONL file.")
    hb_parser.add_argument(
        "--output",
        help="Output JSON file. Default: <input_stem>_scored.json in the same folder.",
    )
    hb_parser.add_argument(
        "--max-workers",
        type=int,
        default=8,
        help="Parallel rubric scoring workers (default: 8).",
    )
    hb_parser.add_argument(
        "--model",
        default=MODEL,
        help=f"Single grading model when panel is disabled (default: {MODEL}).",
    )
    hb_parser.add_argument(
        "--model-registry",
        default=DEFAULT_MODEL_REGISTRY_PATH,
        help=f'Optional model registry JSON (default: "{DEFAULT_MODEL_REGISTRY_PATH}").',
    )
    hb_parser.add_argument(
        "--panel-models",
        default=",".join(DEFAULT_HEALTHBENCH_PANEL_MODELS),
        help=(
            "Comma-separated grading panel model_ids. "
            "Default is Opus 4.6 + Gemini 3.1 Pro Preview + GPT-5.2."
        ),
    )
    hb_parser.add_argument(
        "--disable-panel",
        action="store_true",
        help="Disable panel grading and use only --model.",
    )

    matrix_status_parser = subparsers.add_parser(
        "matrix-status",
        help="Show row/missing status for each model in the 5x2 runs matrix.",
    )
    matrix_status_parser.add_argument(
        "--runs-root",
        default="data/runs",
        help='Runs directory root (default: "data/runs").',
    )
    matrix_status_parser.add_argument(
        "--models",
        default=",".join(DEFAULT_MATRIX_MODELS),
        help="Comma-separated model IDs.",
    )

    eval_matrix_parser = subparsers.add_parser(
        "evaluate-matrix",
        help="Run MedQA and HealthBench evaluation for each model in the matrix.",
    )
    eval_matrix_parser.add_argument(
        "--runs-root",
        default="data/runs",
        help='Runs directory root (default: "data/runs").',
    )
    eval_matrix_parser.add_argument(
        "--models",
        default=",".join(DEFAULT_MATRIX_MODELS),
        help="Comma-separated model IDs.",
    )
    eval_matrix_parser.add_argument(
        "--medqa-model",
        default=MODEL,
        help=f"OpenAI model used for MedQA answer extraction (default: {MODEL}).",
    )
    eval_matrix_parser.add_argument(
        "--healthbench-model",
        default=MODEL,
        help=f"Single grading model used when HealthBench panel is disabled (default: {MODEL}).",
    )
    eval_matrix_parser.add_argument(
        "--healthbench-model-registry",
        default=DEFAULT_MODEL_REGISTRY_PATH,
        help=f'Optional model registry JSON (default: "{DEFAULT_MODEL_REGISTRY_PATH}").',
    )
    eval_matrix_parser.add_argument(
        "--healthbench-panel-models",
        default=",".join(DEFAULT_HEALTHBENCH_PANEL_MODELS),
        help=(
            "Comma-separated grading panel model_ids for HealthBench. "
            "Default is Opus 4.6 + Gemini 3.1 Pro Preview + GPT-5.2."
        ),
    )
    eval_matrix_parser.add_argument(
        "--disable-healthbench-panel",
        action="store_true",
        help="Disable HealthBench panel grading and use --healthbench-model only.",
    )
    eval_matrix_parser.add_argument(
        "--healthbench-workers",
        type=int,
        default=8,
        help="Parallel workers for HealthBench scoring (default: 8).",
    )

    return parser.parse_args(argv)


def main(argv: list[str] | None = None) -> None:
    args = parse_args(argv)
    if args.task == "prepare-healthbench":
        prepare_healthbench_questions(
            dataset_path=args.dataset,
            output_path=args.output,
            limit=args.limit,
            seed=args.seed,
        )
        return

    if args.task == "generate":
        generate_answers(
            input_path=args.input,
            output_path=args.output,
            model=args.model,
            model_id=args.model_id,
            model_registry=args.model_registry,
            max_workers=args.max_workers,
            question_key=args.question_key,
            answer_key=args.answer_key,
            id_key=args.id_key,
            temperature=args.temperature,
            max_tokens=args.max_tokens,
            top_p=args.top_p,
            generation_seed=args.generation_seed,
            enable_search=args.enable_search,
            limit=args.limit,
            resume=args.resume,
            fail_on_error=not args.allow_error_rows,
            checkpoint_every=args.checkpoint_every,
        )
        return

    if args.task == "generate-benchmark-csv":
        generate_benchmark_csv(
            input_path=args.input,
            output_path=args.output,
            benchmark_name=args.benchmark_name,
            runs_root=args.runs_root,
            model_registry=args.model_registry,
            max_workers=args.max_workers,
            question_key=args.question_key,
            answer_key=args.answer_key,
            id_key=args.id_key,
            temperature=args.temperature,
            max_tokens=args.max_tokens,
            top_p=args.top_p,
            generation_seed=args.generation_seed,
            enable_search=args.enable_search,
            limit=args.limit,
            resume=args.resume,
            fail_on_error=not args.allow_error_rows,
            checkpoint_every=args.checkpoint_every,
        )
        return

    if args.task == "matrix-status":
        models = [item.strip() for item in args.models.split(",") if item.strip()]
        matrix_status(runs_root=args.runs_root, models=models)
        return

    if args.task == "evaluate-matrix":
        models = [item.strip() for item in args.models.split(",") if item.strip()]
        healthbench_panel_models = (
            []
            if args.disable_healthbench_panel
            else _parse_model_list(args.healthbench_panel_models)
        )
        evaluate_matrix(
            runs_root=args.runs_root,
            models=models,
            medqa_model=args.medqa_model,
            healthbench_model=args.healthbench_model,
            healthbench_panel_models=healthbench_panel_models or None,
            healthbench_model_registry=args.healthbench_model_registry,
            healthbench_workers=args.healthbench_workers,
        )
        return

    if args.task == "medqa":
        input_path = pathlib.Path(args.input)
        output_path = pathlib.Path(args.output) if args.output else _derive_output(
            input_path, "_processed.jsonl"
        )
        process_medqa(input_path, output_path, model=args.model)
        return

    panel_models = [] if args.disable_panel else _parse_model_list(args.panel_models)
    score_healthbench(
        input_path=args.input,
        max_workers=args.max_workers,
        model=args.model,
        output_path=args.output,
        panel_models=panel_models or None,
        model_registry=args.model_registry,
    )


if __name__ == "__main__":
    main()
